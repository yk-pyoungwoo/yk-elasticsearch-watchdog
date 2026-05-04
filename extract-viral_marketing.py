#!/usr/bin/env python3
# -*- coding: utf-8 -*-
from __future__ import annotations

import os
import sys
import json
import ssl
import time
import base64
import argparse
import zipfile
from datetime import datetime, timezone, timedelta
from urllib.parse import urlencode, urlparse, parse_qs, unquote
from urllib.request import Request, urlopen
from urllib.error import HTTPError, URLError

try:
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter
except ModuleNotFoundError as _e:
    print(
        "extract-viral_marketing.py needs the 'openpyxl' package. "
        "From the repo root run:  python3 -m pip install -r requirements.txt",
        file=sys.stderr,
    )
    raise SystemExit(1) from _e

ES_URL = os.environ.get("ES_URL", "http://localhost:19200").rstrip("/")
ES_USER = os.environ.get("ES_USER", "")
ES_PASS = os.environ.get("ES_PASS", "")
INSECURE_TLS = os.environ.get("INSECURE_TLS", "1") == "1"

SITES = ["brand", "crime", "civil", "divorce", "assault", "drug", "inherit", "traffic", "school"]

BATCH_SIZE = int(os.environ.get("BATCH_SIZE", "2000"))
SLEEP_SEC = float(os.environ.get("SLEEP_SEC", "0"))

MAX_RETRY = int(os.environ.get("MAX_RETRY", "10"))
BACKOFF_MAX_SEC = int(os.environ.get("BACKOFF_MAX_SEC", "30"))
RETRY_STATUS = {429, 502, 503, 504}

OUT_DIR = os.environ.get("OUT_DIR", ".")
OUT_PREFIX = os.environ.get("OUT_PREFIX", "viral_marketing_logs")
CHECKPOINT_DIR = os.environ.get("CHECKPOINT_DIR", ".")

KAKAO_HOSTS = [
    "kakao-blog.yklawfirm.co.kr",
    "kakao-cafe.yklawfirm.co.kr",
    "kakao-daangn.yklawfirm.co.kr",
]

RESERVE_URL = "https://www.yklawfirm.co.kr/counsel/reserve"
INTERNAL_DOMAIN = "yklawfirm.co.kr"

KST = timezone(timedelta(hours=9))

REMOVE_FIELDS = {
    "is_active",
    "channel",
    "title",
    "version",
    "has_been_active",
    "webdriver",
    "tel",
    "is_bot",
    "business_name_1",
    "business_code_1",
    "action",
    "venue",
    "location",
    "asset",
    "seq",
}

TIME_FIELD_KEYWORDS = (
    "timestamp",
    "time",
    "date",
    "first_touch",
    "last_touch",
    "created_at",
    "updated_at",
    "requested_at",
    "occurred_at",
)


def parse_args():
    parser = argparse.ArgumentParser(
        description="Export filtered ES events per day (KST), convert NDJSON to XLSX, zip XLSX files, and remove temporary files."
    )
    parser.add_argument("--start-date", required=True, help='조회 시작 날짜 (KST 기준, 예: "2026-03-02")')
    parser.add_argument("--end-date", required=True, help='조회 종료 날짜 (KST 기준, 예: "2026-03-08")')
    parser.add_argument("--out-dir", default=OUT_DIR, help=f"출력 디렉토리 (기본값: {OUT_DIR})")
    parser.add_argument("--checkpoint-dir", default=CHECKPOINT_DIR, help=f"체크포인트 디렉토리 (기본값: {CHECKPOINT_DIR})")
    parser.add_argument("--out-prefix", default=OUT_PREFIX, help=f"출력 파일 prefix (기본값: {OUT_PREFIX})")
    parser.add_argument(
        "--zip-name",
        default="",
        help='최종 zip 파일명. 미입력 시 예: viral_marketing_logs_2026-03-02~2026-03-08.zip',
    )
    return parser.parse_args()


def parse_kst_date(value: str) -> datetime:
    dt = datetime.strptime(value.strip(), "%Y-%m-%d")
    return dt.replace(tzinfo=KST)


def to_utc(dt: datetime) -> datetime:
    return dt.astimezone(timezone.utc)


def iso_z(dt: datetime) -> str:
    return dt.astimezone(timezone.utc).isoformat(timespec="milliseconds").replace("+00:00", "Z")


def now_kst() -> datetime:
    return datetime.now(KST)


def now_kst_str() -> str:
    return now_kst().strftime("%Y-%m-%d %H:%M:%S")


def iter_months(start_dt_utc: datetime, end_dt_utc: datetime):
    y = start_dt_utc.year
    m = start_dt_utc.month
    cur = datetime(y, m, 1, tzinfo=timezone.utc)

    while cur < end_dt_utc:
        yield cur.year, cur.month
        if cur.month == 12:
            cur = datetime(cur.year + 1, 1, 1, tzinfo=timezone.utc)
        else:
            cur = datetime(cur.year, cur.month + 1, 1, tzinfo=timezone.utc)


def iter_days_inclusive_kst(start_kst: datetime, end_kst: datetime):
    cur = start_kst
    while cur <= end_kst:
        yield cur
        cur += timedelta(days=1)


def build_index(start_dt_utc: datetime, end_dt_utc: datetime) -> str:
    """
    UTC 범위에 걸치는 월(YYYY-MM)을 계산한 뒤,
    각 site별 인덱스 패턴을 만든다.

    기존:
      site-YYYY-MM-* 만 조회
      -> 일별 인덱스(site-YYYY-MM-DD)는 잡히지만
         월별 통합 인덱스(site-YYYY-MM)는 못 잡음

    수정:
      site-YYYY-MM* 로 조회
      -> 아래 둘 다 대응
         - 일별 인덱스: brand-2026-03-01
         - 월별 인덱스: brand-2026-03
    """
    patterns = []
    for year, month in iter_months(start_dt_utc, end_dt_utc):
        ym = f"{year}-{month:02d}"
        for site in SITES:
            patterns.append(f"{site}-{ym}*")
    return ",".join(patterns)


def ssl_ctx():
    if not ES_URL.lower().startswith("https://"):
        return None

    if INSECURE_TLS:
        ctx = ssl.create_default_context()
        ctx.check_hostname = False
        ctx.verify_mode = ssl.CERT_NONE
        return ctx

    return ssl.create_default_context()


def auth_header():
    if ES_USER and ES_PASS:
        token = base64.b64encode(f"{ES_USER}:{ES_PASS}".encode()).decode()
        return {"Authorization": f"Basic {token}"}
    return {}


def load_checkpoint(checkpoint_file):
    if not os.path.exists(checkpoint_file):
        return None
    try:
        with open(checkpoint_file, "r", encoding="utf-8") as f:
            return json.load(f).get("search_after")
    except Exception:
        return None


def save_checkpoint(checkpoint_file, search_after, saved_count):
    with open(checkpoint_file, "w", encoding="utf-8") as f:
        json.dump(
            {
                "search_after": search_after,
                "saved_count": saved_count,
                "updated_at": now_kst_str(),
            },
            f,
            ensure_ascii=False,
            indent=2,
        )


def remove_checkpoint(checkpoint_file):
    if os.path.exists(checkpoint_file):
        os.remove(checkpoint_file)


def http_json(method: str, path: str, body=None, query=None):
    url = ES_URL + path
    if query:
        url += "?" + urlencode(query)

    headers = {"Content-Type": "application/json"}
    headers.update(auth_header())
    data = None if body is None else json.dumps(body).encode("utf-8")
    req = Request(url, data=data, headers=headers, method=method)

    last_err = None
    for attempt in range(1, MAX_RETRY + 1):
        try:
            with urlopen(req, context=ssl_ctx(), timeout=60) as resp:
                raw = resp.read().decode("utf-8", errors="replace")
                return json.loads(raw)

        except HTTPError as e:
            last_err = e
            code = e.code
            err_body = ""
            try:
                err_body = e.read().decode("utf-8", errors="replace")
            except Exception:
                pass

            if code in RETRY_STATUS:
                backoff = min(2 ** (attempt - 1), BACKOFF_MAX_SEC)
                print(f"[WARN] HTTP {code} retry {attempt}/{MAX_RETRY} backoff={backoff}s", file=sys.stderr)
                if err_body:
                    print(f"[WARN] body={err_body[:200]}", file=sys.stderr)
                time.sleep(backoff)
                continue

            print(f"[ERROR] HTTP {code}\n{err_body}", file=sys.stderr)
            raise

        except (URLError, TimeoutError) as e:
            last_err = e
            backoff = min(2 ** (attempt - 1), BACKOFF_MAX_SEC)
            print(f"[WARN] net/timeout retry {attempt}/{MAX_RETRY} backoff={backoff}s: {e}", file=sys.stderr)
            time.sleep(backoff)

    raise RuntimeError(f"Failed after retries. last_err={last_err}")


def field_variants(field_name: str):
    return [field_name, f"{field_name}.keyword"]


def prefix_any(field_name: str, value_prefix: str):
    return [{"prefix": {fv: value_prefix}} for fv in field_variants(field_name)]


def term_any(field_name: str, value: str):
    return [{"term": {fv: value}} for fv in field_variants(field_name)]


def wildcard_any(field_name: str, pattern: str):
    return [{"wildcard": {fv: pattern}} for fv in field_variants(field_name)]


def build_query(time_gte: str, time_lt: str, search_after=None):
    kakao_should = []
    for host in KAKAO_HOSTS:
        kakao_should += prefix_any("url", f"https://{host}/")
        kakao_should += prefix_any("channel", f"https://{host}/")
        kakao_should += wildcard_any("url", f"https://{host}/*")
        kakao_should += wildcard_any("channel", f"https://{host}/*")

    reserve_should = []
    reserve_should += term_any("url", RESERVE_URL)
    reserve_should += term_any("channel", RESERVE_URL)
    reserve_should += term_any("landing_url", RESERVE_URL)
    reserve_should += term_any("page", RESERVE_URL)
    reserve_should += term_any("path", "/counsel/reserve")

    must_not_internal_ref = wildcard_any("referrer", f"*{INTERNAL_DOMAIN}*")

    body = {
        "size": BATCH_SIZE,
        "sort": ["_doc"],
        "_source": True,
        "query": {
            "bool": {
                "filter": [
                    {"range": {"@timestamp": {"gte": time_gte, "lt": time_lt}}}
                ],
                "should": [
                    {"bool": {"should": kakao_should, "minimum_should_match": 1}},
                    {
                        "bool": {
                            "filter": [
                                {"bool": {"should": reserve_should, "minimum_should_match": 1}}
                            ],
                            "must_not": must_not_internal_ref,
                        }
                    },
                ],
                "minimum_should_match": 1,
            }
        },
    }

    if search_after is not None:
        body["search_after"] = search_after

    return body


def parse_from_source(url: str | None):
    if not url:
        return (None, None)
    try:
        qs = parse_qs(urlparse(url).query)
        from_val = qs.get("from", [None])[0]
        source_val = qs.get("source", [None])[0]
        return (
            unquote(from_val) if from_val is not None else None,
            unquote(source_val) if source_val is not None else None,
        )
    except Exception:
        return (None, None)


def netloc(url: str | None):
    if not url:
        return None
    try:
        return urlparse(url).netloc
    except Exception:
        return None


def safe_unquote_string(value: str) -> str:
    cur = value
    for _ in range(3):
        try:
            new_val = unquote(cur)
        except Exception:
            return cur
        if new_val == cur:
            return cur
        cur = new_val
    return cur


def decode_value(value):
    if isinstance(value, str):
        return safe_unquote_string(value)
    if isinstance(value, list):
        return [decode_value(v) for v in value]
    if isinstance(value, dict):
        return {k: decode_value(v) for k, v in value.items()}
    return value


def clean_string_whitespace(value: str) -> str:
    if not isinstance(value, str):
        return value
    return value.replace("\r\n", "\n").replace("\r", "\n")


def normalize_value_for_output(value):
    value = decode_value(value)
    if isinstance(value, str):
        value = clean_string_whitespace(value)
    return value


def looks_like_datetime_field(field_name: str) -> bool:
    if not field_name:
        return False
    lower = field_name.lower()
    return any(keyword in lower for keyword in TIME_FIELD_KEYWORDS)


def parse_datetime_string(value: str):
    if not isinstance(value, str):
        return None

    s = value.strip()
    if not s:
        return None

    try:
        if s.endswith("Z"):
            s = s[:-1] + "+00:00"
        return datetime.fromisoformat(s)
    except Exception:
        return None


def convert_datetime_to_kst_for_excel(value: str):
    """
    규칙:
    - 이미 +09:00 이면 그대로 둠
    - UTC/Z/기타 timezone이면 KST로 변환
    - timezone 없는 naive 값이면 이미 한국시간으로 간주하여 그대로 둠

    반환:
    - datetime 객체 (timezone 제거된 naive datetime)
    - 변환 불가 시 원래 값
    """
    dt = parse_datetime_string(value)
    if dt is None:
        return value

    if dt.tzinfo is None:
        return dt

    offset = dt.utcoffset()
    if offset == timedelta(hours=9):
        return dt.replace(tzinfo=None)

    return dt.astimezone(KST).replace(tzinfo=None)


def clean_doc(doc: dict) -> dict:
    doc = decode_value(doc)

    url_value = doc.get("url") or doc.get("channel")
    host = netloc(url_value)
    if host in KAKAO_HOSTS:
        parsed_from, parsed_source = parse_from_source(url_value)
        if doc.get("from") is None and parsed_from is not None:
            doc["from"] = parsed_from
        if doc.get("source") is None and parsed_source is not None:
            doc["source"] = parsed_source

    for field in REMOVE_FIELDS:
        doc.pop(field, None)

    doc = {k: normalize_value_for_output(v) for k, v in doc.items()}
    return doc


def flatten_value(value, field_name=""):
    value = normalize_value_for_output(value)

    if isinstance(value, str) and looks_like_datetime_field(field_name):
        value = convert_datetime_to_kst_for_excel(value)

    if isinstance(value, dict):
        decoded_dict = decode_value(value)
        return json.dumps(decoded_dict, ensure_ascii=False)

    if isinstance(value, list):
        decoded_list = decode_value(value)
        return json.dumps(decoded_list, ensure_ascii=False)

    return value


def flatten_dict(data, parent_key="", sep="."):
    items = {}
    for key, value in data.items():
        new_key = f"{parent_key}{sep}{key}" if parent_key else key
        if isinstance(value, dict):
            items.update(flatten_dict(value, new_key, sep=sep))
        else:
            items[new_key] = flatten_value(value, new_key)
    return items


def ndjson_to_xlsx(ndjson_path: str, xlsx_path: str):
    rows = []
    headers = []

    with open(ndjson_path, "r", encoding="utf-8") as f:
        for line_no, line in enumerate(f, start=1):
            line = line.strip()
            if not line:
                continue

            try:
                obj = json.loads(line)
            except Exception as e:
                raise ValueError(f"NDJSON parse error: {ndjson_path}:{line_no}: {e}") from e

            obj = decode_value(obj)
            flat = flatten_dict(obj)
            rows.append(flat)

            for key in flat.keys():
                if key not in headers:
                    headers.append(key)

    wb = Workbook()
    ws = wb.active
    ws.title = "viral_marketing"

    if not headers:
        headers = ["message"]
        ws.append(headers)
        ws.append(["no data"])
    else:
        ws.append(headers)
        for row in rows:
            ws.append([row.get(col, "") for col in headers])

        for row_idx in range(2, ws.max_row + 1):
            for col_idx in range(1, ws.max_column + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                if isinstance(cell.value, datetime):
                    cell.number_format = "yyyy-mm-dd hh:mm:ss"

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    for idx, col_name in enumerate(headers, start=1):
        max_len = len(str(col_name))
        for row_idx in range(2, ws.max_row + 1):
            val = ws.cell(row=row_idx, column=idx).value
            if val is None:
                continue

            if isinstance(val, datetime):
                text_val = val.strftime("%Y-%m-%d %H:%M:%S")
            else:
                text_val = str(val)

            max_len = max(max_len, len(text_val))

        ws.column_dimensions[get_column_letter(idx)].width = min(max_len + 2, 60)

    wb.save(xlsx_path)


def create_zip(zip_path: str, file_paths: list[str]):
    with zipfile.ZipFile(zip_path, "w", compression=zipfile.ZIP_DEFLATED) as zf:
        for file_path in file_paths:
            zf.write(file_path, arcname=os.path.basename(file_path))


def export_one_day(day_kst: datetime, out_dir: str, checkpoint_dir: str, out_prefix: str) -> str:
    day_str = day_kst.strftime("%Y-%m-%d")
    next_day_kst = day_kst + timedelta(days=1)

    start_utc = to_utc(day_kst)
    end_utc = to_utc(next_day_kst)

    time_gte = iso_z(start_utc)
    time_lt = iso_z(end_utc)
    index = build_index(start_utc, end_utc)

    os.makedirs(out_dir, exist_ok=True)
    os.makedirs(checkpoint_dir, exist_ok=True)

    ndjson_file = os.path.join(out_dir, f"{out_prefix}_{day_str}.ndjson")
    xlsx_file = os.path.join(out_dir, f"{out_prefix}_{day_str}.xlsx")
    checkpoint_file = os.path.join(checkpoint_dir, f"checkpoint_{day_str}.json")

    print("=" * 80)
    print(f"[INFO] KST_DAY={day_str}")
    print(f"[INFO] KST_RANGE={day_kst.isoformat()} ~ {next_day_kst.isoformat()}")
    print(f"[INFO] UTC_RANGE={time_gte} ~ {time_lt}")
    print(f"[INFO] INDEX={index}")
    print(f"[INFO] NDJSON_FILE={ndjson_file}")
    print(f"[INFO] XLSX_FILE={xlsx_file}")
    print(f"[INFO] CHECKPOINT={checkpoint_file}")

    search_after = load_checkpoint(checkpoint_file)
    if search_after:
        print(f"[INFO] resume search_after={search_after}")

    saved = 0
    open_mode = "a" if search_after else "w"

    with open(ndjson_file, open_mode, encoding="utf-8") as out:
        while True:
            body = build_query(time_gte, time_lt, search_after)
            res = http_json("POST", f"/{index}/_search", body=body)

            hits = (res.get("hits") or {}).get("hits") or []
            if not hits:
                print(f"[DONE] {day_str} no more hits")
                break

            for hit in hits:
                doc = hit.get("_source") or {}
                cleaned = clean_doc(doc)
                out.write(json.dumps(cleaned, ensure_ascii=False) + "\n")
                saved += 1

            search_after = hits[-1].get("sort")
            save_checkpoint(checkpoint_file, search_after, saved)

            print(f"[INFO] {day_str} batch ok. total saved={saved}")

            if SLEEP_SEC > 0:
                time.sleep(SLEEP_SEC)

    remove_checkpoint(checkpoint_file)
    print(f"[INFO] removed checkpoint: {checkpoint_file}")

    print(f"[INFO] converting NDJSON -> XLSX: {ndjson_file} -> {xlsx_file}")
    ndjson_to_xlsx(ndjson_file, xlsx_file)

    if not os.path.exists(xlsx_file):
        raise RuntimeError(f"XLSX conversion failed: {xlsx_file}")

    if os.path.exists(ndjson_file):
        os.remove(ndjson_file)
        print(f"[INFO] removed ndjson after xlsx conversion: {ndjson_file}")

    print(f"[FINISH] {day_str} saved={saved}, xlsx={xlsx_file}")
    return xlsx_file


def main():
    args = parse_args()

    start_kst = parse_kst_date(args.start_date)
    end_kst = parse_kst_date(args.end_date)

    if start_kst > end_kst:
        raise ValueError("start-date 는 end-date 보다 과거이거나 같아야 합니다.")

    os.makedirs(args.out_dir, exist_ok=True)
    os.makedirs(args.checkpoint_dir, exist_ok=True)

    zip_name = args.zip_name.strip() or f"viral_marketing_logs_{args.start_date}~{args.end_date}.zip"
    zip_path = os.path.join(args.out_dir, zip_name)

    print(f"[INFO] ES_URL={ES_URL}")
    print(f"[INFO] BATCH_SIZE={BATCH_SIZE}, SLEEP_SEC={SLEEP_SEC}")
    print(f"[INFO] START_DATE(KST)={args.start_date}")
    print(f"[INFO] END_DATE(KST)={args.end_date}")
    print(f"[INFO] OUT_DIR={args.out_dir}")
    print(f"[INFO] CHECKPOINT_DIR={args.checkpoint_dir}")
    print(f"[INFO] OUT_PREFIX={args.out_prefix}")
    print(f"[INFO] ZIP_PATH={zip_path}")

    created_xlsx_files = []

    for day_kst in iter_days_inclusive_kst(start_kst, end_kst):
        xlsx_file = export_one_day(day_kst, args.out_dir, args.checkpoint_dir, args.out_prefix)
        created_xlsx_files.append(xlsx_file)

    print(f"[INFO] creating zip: {zip_path}")
    create_zip(zip_path, created_xlsx_files)

    if not os.path.exists(zip_path):
        raise RuntimeError(f"ZIP creation failed: {zip_path}")

    print(f"[INFO] zip created: {zip_path}")

    for xlsx_file in created_xlsx_files:
        if os.path.exists(xlsx_file):
            os.remove(xlsx_file)
            print(f"[INFO] removed xlsx after zip creation: {xlsx_file}")

    print(f"[FINISH] zip created and temporary xlsx files removed: {zip_path}")


if __name__ == "__main__":
    main()